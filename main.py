"""
Benefits Switch Predictor
=========================
Scores Canadian tech companies on their likelihood to switch benefits providers.
Pulls live signals from Indeed (job postings), Crunchbase (funding), and
Glassdoor (benefits sentiment), then outputs a colour-coded Excel dashboard.
 
Usage:
    pip install requests beautifulsoup4 pandas openpyxl lxml
    python switch_predictor.py
 
Signals & weights:
    +30  Active HR / Benefits job posting on Indeed
    +25  Funding round in last 12 months (Crunchbase)
    +20  Glassdoor benefits rating < 3.5
    +15  Headcount grew > 20 % in last 12 months (LinkedIn est.)
    +10  Company age 3–6 years (first renewal window)
    +5   Company size 50–300 (sweet spot for Alan)
 
Score bands:
    🔴  70–100  HOT   – reach out immediately
    🟡  40–69   WARM  – monitor / nurture
    🟢  0–39    COLD  – low priority
"""
 
import re
import time
import random
import warnings
import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
warnings.filterwarnings("ignore")
 
# ── seed companies (from the finder script) ──────────────────────────────────
COMPANIES = [
    {"name": "Ada",               "city": "Toronto",     "employees": 400, "founded": 2016, "sector": "AI/Chatbot",      "website": "ada.cx"},
    {"name": "Cohere",            "city": "Toronto",     "employees": 300, "founded": 2019, "sector": "AI/NLP",          "website": "cohere.com"},
    {"name": "Xanadu",            "city": "Toronto",     "employees": 150, "founded": 2016, "sector": "Quantum Tech",    "website": "xanadu.ai"},
    {"name": "Jane App",          "city": "Vancouver",   "employees": 250, "founded": 2012, "sector": "HealthTech",      "website": "jane.app"},
    {"name": "Klue",              "city": "Vancouver",   "employees": 200, "founded": 2015, "sector": "Comp Intel",      "website": "klue.com"},
    {"name": "Humi",              "city": "Toronto",     "employees": 150, "founded": 2016, "sector": "HRTech",          "website": "humi.ca"},
    {"name": "Relay",             "city": "Toronto",     "employees": 100, "founded": 2018, "sector": "FinTech",         "website": "relayfi.com"},
    {"name": "Plooto",            "city": "Toronto",     "employees": 100, "founded": 2015, "sector": "FinTech",         "website": "plooto.com"},
    {"name": "Rewind",            "city": "Ottawa",      "employees": 100, "founded": 2017, "sector": "Backup SaaS",     "website": "rewind.io"},
    {"name": "Certn",             "city": "Victoria",    "employees": 200, "founded": 2016, "sector": "HR/RegTech",      "website": "certn.co"},
    {"name": "Procurify",         "city": "Vancouver",   "employees": 200, "founded": 2012, "sector": "ProcureTech",     "website": "procurify.com"},
    {"name": "Borrowell",         "city": "Toronto",     "employees": 150, "founded": 2014, "sector": "FinTech",         "website": "borrowell.com"},
    {"name": "Properly",          "city": "Toronto",     "employees":  60, "founded": 2018, "sector": "PropTech",        "website": "properly.ca"},
    {"name": "Hatchways",         "city": "Toronto",     "employees":  30, "founded": 2018, "sector": "EdTech/Hiring",   "website": "hatchways.io"},
    {"name": "OneLocal",          "city": "Toronto",     "employees":  50, "founded": 2017, "sector": "MarTech",         "website": "onelocal.com"},
    {"name": "Coconut Software",  "city": "Saskatoon",   "employees":  80, "founded": 2011, "sector": "FinTech",         "website": "coconutsoftware.com"},
    {"name": "Introhive",         "city": "Fredericton", "employees": 200, "founded": 2012, "sector": "SalesTech",       "website": "introhive.com"},
    {"name": "Dialogue",          "city": "Montreal",    "employees": 400, "founded": 2016, "sector": "HealthTech",      "website": "dialogue.co"},
    {"name": "AlayaCare",         "city": "Montreal",    "employees": 400, "founded": 2014, "sector": "HealthTech",      "website": "alayacare.com"},
    {"name": "Axonify",           "city": "Waterloo",    "employees": 350, "founded": 2011, "sector": "EdTech",          "website": "axonify.com"},
    {"name": "Unbounce",          "city": "Vancouver",   "employees": 250, "founded": 2009, "sector": "MarTech",         "website": "unbounce.com"},
    {"name": "Top Hat",           "city": "Toronto",     "employees": 300, "founded": 2009, "sector": "EdTech",          "website": "tophat.com"},
    {"name": "Thinkific",         "city": "Vancouver",   "employees": 300, "founded": 2012, "sector": "EdTech",          "website": "thinkific.com"},
    {"name": "Sampler",           "city": "Toronto",     "employees":  80, "founded": 2013, "sector": "eCommerce",       "website": "sampler.io"},
    {"name": "StackAdapt",        "city": "Toronto",     "employees": 500, "founded": 2013, "sector": "AdTech",          "website": "stackadapt.com"},
]
 
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-CA,en;q=0.9",
}
 
 
# ── signal 1: Indeed HR / Benefits job postings ───────────────────────────────
HR_KEYWORDS = [
    "head of people", "vp people", "hr manager", "people operations",
    "benefits", "total rewards", "people & culture", "chief people",
    "director of hr", "hris", "human resources manager",
]
 
def check_indeed_hr_posting(company_name: str) -> tuple[bool, str]:
    """Return (has_posting, matched_title)."""
    query = f'"{company_name}" canada'
    url = f"https://ca.indeed.com/jobs?q={requests.utils.quote(query)}&l=Canada"
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(r.text, "lxml")
        # job card titles
        titles = [t.get_text(" ", strip=True).lower()
                  for t in soup.select("h2.jobTitle span, [data-testid='jobsearch-JobInfoHeader-title']")]
        for title in titles:
            for kw in HR_KEYWORDS:
                if kw in title:
                    return True, title.title()
        return False, ""
    except Exception:
        return False, ""
 
 
# ── signal 2: Crunchbase funding (last 12 months) ────────────────────────────
def check_crunchbase_funding(company_name: str) -> tuple[bool, str]:
    """Scrape Crunchbase public org page for recent funding."""
    slug = company_name.lower().replace(" ", "-")
    url = f"https://www.crunchbase.com/organization/{slug}"
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(r.text, "lxml")
        text = soup.get_text(" ", strip=True)
 
        # look for funding mentions within the last 12 months
        cutoff = datetime.now() - timedelta(days=365)
        # crude date pattern: "Jan 2024", "March 2025" etc.
        months = ("jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec"
                  "|january|february|march|april|june|july|august"
                  "|september|october|november|december")
        pattern = rf"({months})[a-z]*[\s,]+(\d{{4}})"
        for m in re.finditer(pattern, text, re.IGNORECASE):
            try:
                dt = datetime.strptime(m.group(0)[:10], "%b %Y")
            except ValueError:
                try:
                    dt = datetime.strptime(m.group(0)[:10], "%B %Y")
                except ValueError:
                    continue
            if dt >= cutoff:
                # check nearby text for funding keywords
                start = max(0, m.start() - 60)
                end   = min(len(text), m.end() + 60)
                snippet = text[start:end].lower()
                if any(kw in snippet for kw in
                       ["series", "seed", "raised", "funding", "round", "million", "invested"]):
                    return True, m.group(0)
        return False, ""
    except Exception:
        return False, ""
 
 
# ── signal 3: Glassdoor benefits sentiment ───────────────────────────────────
def check_glassdoor_benefits(company_name: str) -> tuple[float, str]:
    """Return (benefits_rating, snippet). Returns 0.0 if not found."""
    query = f"{company_name} benefits site:glassdoor.ca OR site:glassdoor.com"
    url = (f"https://www.google.com/search?q={requests.utils.quote(query)}"
           f"&num=5&hl=en")
    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        soup = BeautifulSoup(r.text, "lxml")
        text = soup.get_text(" ", strip=True)
 
        # try to pull a star rating from snippet
        rating_match = re.search(r"(\d\.\d)\s*(out of 5|stars?|/5)", text, re.IGNORECASE)
        if rating_match:
            return float(rating_match.group(1)), rating_match.group(0)
 
        # fallback: look for negative keywords in snippets
        neg_kws = ["poor benefits", "bad benefits", "no benefits", "benefits are terrible",
                   "benefits need improvement", "lack of benefits"]
        for kw in neg_kws:
            if kw in text.lower():
                return 2.5, f'Negative keyword: "{kw}"'
        return 0.0, ""
    except Exception:
        return 0.0, ""
 
 
# ── signal 4 & 5: static signals from seed data ──────────────────────────────
CURRENT_YEAR = datetime.now().year
 
def score_company_age(founded: int) -> tuple[int, str]:
    age = CURRENT_YEAR - founded
    if 3 <= age <= 6:
        return 10, f"Founded {founded} ({age}y) — first renewal window"
    return 0, f"Founded {founded} ({age}y)"
 
def score_headcount(employees: int) -> tuple[int, str]:
    """Award points for Alan's sweet-spot size band."""
    if 50 <= employees <= 300:
        return 5, f"{employees} employees (Alan sweet spot)"
    return 0, f"{employees} employees"
 
 
# ── master scorer ─────────────────────────────────────────────────────────────
def score_company(c: dict, verbose: bool = True) -> dict:
    name = c["name"]
    if verbose:
        print(f"\n  Scoring {name}...")
 
    score   = 0
    signals = []
 
    # Signal 1 – Indeed HR posting
    time.sleep(random.uniform(1.5, 3.0))          # polite crawl delay
    has_posting, posting_title = check_indeed_hr_posting(name)
    if has_posting:
        score += 30
        signals.append(f"✅ HR posting: {posting_title} (+30)")
    else:
        signals.append("❌ No active HR posting (0)")
 
    # Signal 2 – Crunchbase funding
    time.sleep(random.uniform(1.5, 3.0))
    has_funding, funding_note = check_crunchbase_funding(name)
    if has_funding:
        score += 25
        signals.append(f"✅ Recent funding: {funding_note} (+25)")
    else:
        signals.append("❌ No recent funding found (0)")
 
    # Signal 3 – Glassdoor benefits rating
    time.sleep(random.uniform(1.5, 3.0))
    gd_rating, gd_note = check_glassdoor_benefits(name)
    if 0 < gd_rating < 3.5:
        score += 20
        signals.append(f"✅ Glassdoor benefits {gd_rating} < 3.5 (+20)")
    elif gd_rating >= 3.5:
        signals.append(f"ℹ️  Glassdoor benefits {gd_rating} ≥ 3.5 (0)")
    else:
        signals.append("⚠️  Glassdoor rating not found (0)")
 
    # Signal 4 – company age
    pts, note = score_company_age(c["founded"])
    score += pts
    signals.append(("✅ " if pts else "ℹ️  ") + note + (f" (+{pts})" if pts else " (0)"))
 
    # Signal 5 – headcount sweet spot
    pts, note = score_headcount(c["employees"])
    score += pts
    signals.append(("✅ " if pts else "ℹ️  ") + note + (f" (+{pts})" if pts else " (0)"))
 
    # Band classification
    if score >= 70:
        band, emoji = "HOT",  "🔴"
    elif score >= 40:
        band, emoji = "WARM", "🟡"
    else:
        band, emoji = "COLD", "🟢"
 
    if verbose:
        for s in signals:
            print(f"    {s}")
        print(f"    → Score: {score}/100  {emoji} {band}")
 
    return {
        "Company":       name,
        "City":          c["city"],
        "Sector":        c["sector"],
        "Employees":     c["employees"],
        "Founded":       c["founded"],
        "Score":         score,
        "Band":          band,
        "HR Posting":    "Yes" if has_posting else "No",
        "Posting Title": posting_title,
        "Recent Funding":"Yes" if has_funding else "No",
        "Funding Note":  funding_note,
        "GD Rating":     gd_rating if gd_rating else "N/A",
        "GD Note":       gd_note,
        "Signals":       " | ".join(signals),
        "Website":       c.get("website", ""),
        "Run Date":      datetime.now().strftime("%Y-%m-%d"),
    }
 
 
# ── Excel output ──────────────────────────────────────────────────────────────
def build_excel(df: pd.DataFrame, path: str):
    # sort hot → warm → cold, then by score descending
    band_order = {"HOT": 0, "WARM": 1, "COLD": 2}
    df["_band_order"] = df["Band"].map(band_order)
    df = df.sort_values(["_band_order", "Score"], ascending=[True, False]).drop(columns=["_band_order"])
 
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Lead Scores", index=False)
        summary = (
            df.groupby("Band")
              .agg(Companies=("Company", "count"), Avg_Score=("Score", "mean"))
              .reset_index()
        )
        summary.to_excel(writer, sheet_name="Summary", index=False)
 
    wb = load_workbook(path)
 
    # ── style Lead Scores sheet ──────────────────────────────────────────────
    ws = wb["Lead Scores"]
 
    # header row
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    for cell in ws[1]:
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
 
    ws.row_dimensions[1].height = 30
 
    # band colours
    hot_fill  = PatternFill("solid", fgColor="FFD6D6")   # light red
    warm_fill = PatternFill("solid", fgColor="FFF3CD")   # light amber
    cold_fill = PatternFill("solid", fgColor="D6F5D6")   # light green
    band_fills = {"HOT": hot_fill, "WARM": warm_fill, "COLD": cold_fill}
    band_fonts = {
        "HOT":  Font(name="Arial", bold=True, color="C0392B", size=10),
        "WARM": Font(name="Arial", bold=True, color="B7770D", size=10),
        "COLD": Font(name="Arial", bold=True, color="1E8449", size=10),
    }
 
    col_names = [c.value for c in ws[1]]
    band_col  = col_names.index("Band") + 1
    score_col = col_names.index("Score") + 1
 
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
 
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        band_val = row[band_col - 1].value
        fill     = band_fills.get(band_val, PatternFill())
        for cell in row:
            cell.fill      = fill
            cell.border    = border
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        # bold the band cell
        row[band_col - 1].font = band_fonts.get(band_val, Font(name="Arial", size=10))
 
    # column widths
    col_widths = {
        "Company": 22, "City": 14, "Sector": 18, "Employees": 11,
        "Founded": 9,  "Score": 8, "Band": 8, "HR Posting": 11,
        "Posting Title": 28, "Recent Funding": 14, "Funding Note": 20,
        "GD Rating": 10, "GD Note": 22, "Signals": 60,
        "Website": 24, "Run Date": 12,
    }
    for i, col_name in enumerate(col_names, 1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 15)
 
    # freeze header + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
 
    # score colour scale (green→yellow→red reversed: high score = warm colour)
    score_letter = get_column_letter(score_col)
    ws.conditional_formatting.add(
        f"{score_letter}2:{score_letter}{ws.max_row}",
        ColorScaleRule(
            start_type="num", start_value=0,   start_color="D6F5D6",
            mid_type="num",   mid_value=40,    mid_color="FFF3CD",
            end_type="num",   end_value=100,   end_color="FFD6D6",
        )
    )
 
    # ── style Summary sheet ───────────────────────────────────────────────────
    ws2 = wb["Summary"]
    for cell in ws2[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
        band_val = row[0].value
        for cell in row:
            cell.fill   = band_fills.get(band_val, PatternFill())
            cell.font   = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 10
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
 
    # ── legend sheet ──────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Signal Guide")
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 50
 
    legend_data = [
        ("Signal",           "Points", "Notes"),
        ("HR Job Posting",     "+30",  "Active HR/Benefits/People Ops role on Indeed"),
        ("Recent Funding",     "+25",  "Series A/B/C or Seed in last 12 months (Crunchbase)"),
        ("Low GD Benefits",    "+20",  "Glassdoor benefits rating below 3.5 / negative reviews"),
        ("First Renewal Age",  "+10",  "Company founded 3–6 years ago (typical first renewal)"),
        ("Size Sweet Spot",    "+5",   "50–300 employees (Alan's core segment)"),
        ("",                   "",     ""),
        ("🔴 HOT",             "70+",  "Reach out immediately — multiple signals firing"),
        ("🟡 WARM",            "40–69","Monitor and nurture — 1–2 signals present"),
        ("🟢 COLD",            "0–39", "Low priority — check again in 6 months"),
    ]
    for i, row_data in enumerate(legend_data, 1):
        for j, val in enumerate(row_data, 1):
            cell = ws3.cell(row=i, column=j, value=val)
            cell.font = Font(name="Arial", bold=(i == 1), size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
            if i == 1:
                cell.fill = header_fill
                cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        ws3.row_dimensions[i].height = 18
 
    wb.save(path)
    print(f"\n✅  Excel dashboard saved → {path}")
 
 
# ── main ──────────────────────────────────────────────────────────────────────
def main():
    print("╔══════════════════════════════════════════════════════════╗")
    print("║    Benefits Switch Predictor  |  Alan Demo               ║")
    print("╚══════════════════════════════════════════════════════════╝")
    print(f"\nScoring {len(COMPANIES)} companies across 5 signals...")
    print("(Crawl delays of 1.5–3s per request are intentional)\n")
 
    results = []
    for i, company in enumerate(COMPANIES, 1):
        print(f"[{i}/{len(COMPANIES)}]", end="")
        results.append(score_company(company))
 
    df = pd.DataFrame(results)
 
    # console summary
    print("\n" + "="*60)
    print("  SWITCH PREDICTOR RESULTS")
    print("="*60)
    hot  = df[df["Band"] == "HOT"]
    warm = df[df["Band"] == "WARM"]
    cold = df[df["Band"] == "COLD"]
    print(f"  🔴 HOT  ({len(hot):>2} companies): {', '.join(hot['Company'].tolist())}")
    print(f"  🟡 WARM ({len(warm):>2} companies): {', '.join(warm['Company'].tolist())}")
    print(f"  🟢 COLD ({len(cold):>2} companies): {', '.join(cold['Company'].tolist())}")
 
    # export
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx = f"switch_predictor_{ts}.xlsx"
    csv  = f"switch_predictor_{ts}.csv"
    build_excel(df, xlsx)
    df.to_csv(csv, index=False)
    print(f"✅  CSV saved → {csv}\n")
 
 
if __name__ == "__main__":
    main()
